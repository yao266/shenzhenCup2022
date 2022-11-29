#coding:utf-8
from openpyxl import Workbook
from openpyxl import load_workbook
from lxml import etree
import requests
import time,re
import random
import pandas as pd

def get_html(url):
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Cache-Control': 'max-age=0',
        'Connection': 'keep-alive',
        'Cookie': 'SINAGLOBAL=9576599559727.137.1658732154803; wvr=6; SCF=AnqAUJ-xFsNnyiHwaTbnqEjUWT_s_I7XDWWd-Zp4fsU_WzJrKRxrxMz7LISx7KyyC-NJI2hcT-VdmfWQIM4l2Gw.; ALF=1690685606; SUB=_2A25P4JJDDeRhGeBM7FYW8ibIwj-IHXVtKj4LrDV8PUJbkNAKLXfykW1NRNCwEyCOJueoaBnUHZX5LMDHG7xe97wy; SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9Wh-zEgEAzx2JnpAzDb.BIGD5NHD95QceoMXS0zRSh.0Ws4DqcjgwrU0wCH8SCHWSb-4xCH8SCHWSb-4xBtt; UOR=,,www.baidu.com; webim_unReadCount={"time":1659170907665,"dm_pub_total":0,"chat_group_client":0,"chat_group_notice":0,"allcountNum":0,"msgbox":0}; _s_tentry=s.weibo.com; Apache=2404674415337.489.1659228189699; ULV=1659228189722:3:3:1:2404674415337.489.1659228189699:1658884782989',
        'Pragma': 'no-cache',
        'sec-ch-ua': '" Not;A Brand";v="99", "Microsoft Edge";v="103", "Chromium";v="103"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform':'"Windows"',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'none',
        'Sec-Fetch-User': '?1',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.5060.134 Safari/537.36 Edg/103.0.1264.77'
    }
    res = requests.get(url,headers=headers)
    if res.status_code == 200:
        print('获取成功')
        return res.text
    else:
        print('失败')

def jianxi(res):
    data = []
    res = re.findall('<!--card-wrap-->(.*?)<!--/card-wrap-->',res,re.S)
    print(res)
    for r in res:
        xp = etree.HTML(r)
        n = xp.xpath('//p[@class="txt" and @node-type="feed_list_content_full"]//text()')
        if len(n) == 0:
            n = xp.xpath('//p[@class="txt" and @node-type="feed_list_content"]//text()')
        username = xp.xpath('//div[@class="info"]/div/a/text()')[0]
        t = xp.xpath('//div[@class="content"]/p[@class="from"]/a[1]/text()')
        share = xp.xpath('//div[@class="card-act"]//li[1]/a/text()')[0].strip()
        p = xp.xpath('//div[@class="card-act"]//ul//li[2]/a/text()')[0]
        d = xp.xpath('//div[@class="card-act"]//ul//li[3]//a//button/span[2]/text()')
       
        if share=='转发':
            share ='0'

        
        if len(d) != 0:
            d = d[0]
        else :
            d = '0'
        if d =='赞':
            d = '0'
        p =re.findall('\d*',p)
        p = ''.join('%s' % r.split() for r in p).replace('[', '').replace(']', '').replace('\'', '')
        if p == '':
            p = '0'
        t = ''.join(t[0].split())
        n = ''.join('%s' %r.split() for r in n).replace('[','').replace(']','').replace('\'','')
        n = re.sub(r'\\u...','',n)
        n = re.sub(r'收起全文d','',n)
        data.append({'发帖人':username,'时间': t, '评论数': p, '点赞数': d,'转发数':share, '内容': n})
    return data

def write_data(datas):
    wb = load_workbook('易烊千玺.xlsx')
    ws = wb.create_sheet('易烊千玺', 0)
    ys = {
        'A':'发帖人',
        'B':'时间',
        'C':'评论数',
        'D':'点赞数',
        'E':'转发数',
        'F':'内容'
    }
    for key, value in ys.items():
        ws[key + '1'] = value
    b = 0
    for data in datas:
        for n in range(len(list(data.values())[0])):
            for key, value in ys.items():
                ws[key + str(n + 2 + b)] = list(data.values())[0][n][value]
        b += len(list(data.values())[0])
    wb.save('易烊千玺.xlsx')

def get_date(searchdate):
    return str(searchdate).split(' ')[0]+'-'+str(searchdate).split(' ')[1].split(':')[0]

if __name__ == '__main__':
    wb = Workbook()
    wb.save('易烊千玺.xlsx')
    datas = []
    # daterange = pd.date_range(start='2022-07-06 16:00:00',freq='1H',periods=561)
    url = "https://m.weibo.cn/detail/4778846533977535"
    print(url)
    res = get_html(url)
    data = jianxi(res)
    print(data)
    datas.append({data})
    time.sleep(random.uniform(1,4))
    write_data(datas)
